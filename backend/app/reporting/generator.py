"""
Reporting module for generating PDF/Excel reports on OCR performance,
correction statistics, and model training progress.
"""

import io
import json
import logging
from typing import Dict, List, Optional
from datetime import datetime, timedelta
from dataclasses import dataclass

from sqlalchemy.orm import Session
from sqlalchemy import text

logger = logging.getLogger(__name__)


@dataclass
class ReportConfig:
    """Configuration for report generation."""
    title: str
    period_start: datetime
    period_end: datetime
    format: str = "pdf"  # pdf or excel
    include_charts: bool = True
    include_details: bool = True


class ReportGenerator:
    """
    Generates performance reports from OCR correction data.
    Supports PDF and Excel output formats.
    """

    def __init__(self, db: Session):
        self.db = db

    def generate_correction_report(self, config: ReportConfig) -> bytes:
        """
        Generate a comprehensive correction statistics report.
        
        Returns:
            bytes: PDF or Excel file content.
        """
        stats = self._gather_correction_stats(config.period_start, config.period_end)
        training_stats = self._gather_training_stats()
        
        if config.format == "excel":
            return self._generate_excel_report(stats, training_stats, config)
        else:
            return self._generate_pdf_report(stats, training_stats, config)

    def _gather_correction_stats(self, start: datetime, end: datetime) -> Dict:
        """Gather correction statistics from database."""
        result = self.db.execute(text("""
            SELECT
                COUNT(*) as total_regions,
                COUNT(CASE WHEN corrected_text IS NOT NULL THEN 1 END) as corrected_count,
                COUNT(CASE WHEN status = 'gold_standard' THEN 1 END) as gold_count,
                COUNT(CASE WHEN status = 'pending' THEN 1 END) as pending_count,
                AVG(confidence) as avg_confidence,
                COUNT(CASE WHEN is_medical_term = TRUE THEN 1 END) as medical_term_count,
                COUNT(CASE WHEN script_class = 'arabic' THEN 1 END) as arabic_count,
                COUNT(CASE WHEN script_class = 'latin' THEN 1 END) as latin_count,
                COUNT(CASE WHEN script_class = 'mixed' THEN 1 END) as mixed_count,
                COUNT(CASE WHEN confidence < 0.5 THEN 1 END) as low_confidence_count,
                COUNT(CASE WHEN confidence >= 0.9 THEN 1 END) as high_confidence_count
            FROM text_regions
            WHERE created_at BETWEEN :start AND :end
        """), {"start": start, "end": end}).fetchone()

        # Top corrected terms
        top_terms = self.db.execute(text("""
            SELECT corrected_text, COUNT(*) as correction_count
            FROM text_regions
            WHERE corrected_text IS NOT NULL
              AND created_at BETWEEN :start AND :end
            GROUP BY corrected_text
            ORDER BY correction_count DESC
            LIMIT 20
        """), {"start": start, "end": end}).fetchall()

        # Daily correction trend
        daily_trend = self.db.execute(text("""
            SELECT DATE(created_at) as date, COUNT(*) as count
            FROM text_regions
            WHERE corrected_text IS NOT NULL
              AND created_at BETWEEN :start AND :end
            GROUP BY DATE(created_at)
            ORDER BY date
        """), {"start": start, "end": end}).fetchall()

        # User activity
        user_activity = self.db.execute(text("""
            SELECT user_id, COUNT(*) as corrections, AVG(confidence) as avg_confidence
            FROM text_regions
            WHERE corrected_text IS NOT NULL
              AND created_at BETWEEN :start AND :end
            GROUP BY user_id
            ORDER BY corrections DESC
        """), {"start": start, "end": end}).fetchall()

        return {
            "summary": {
                "total_regions": result.total_regions or 0,
                "corrected": result.corrected_count or 0,
                "gold_standard": result.gold_count or 0,
                "pending": result.pending_count or 0,
                "avg_confidence": float(result.avg_confidence or 0),
                "medical_terms": result.medical_term_count or 0,
                "arabic": result.arabic_count or 0,
                "latin": result.latin_count or 0,
                "mixed": result.mixed_count or 0,
                "low_confidence": result.low_confidence_count or 0,
                "high_confidence": result.high_confidence_count or 0,
                "correction_rate": (
                    (result.corrected_count / result.total_regions * 100)
                    if result.total_regions > 0 else 0
                ),
            },
            "top_corrected_terms": [
                {"term": row.corrected_text, "count": row.correction_count}
                for row in top_terms
            ],
            "daily_trend": [
                {"date": str(row.date), "count": row.count}
                for row in daily_trend
            ],
            "user_activity": [
                {
                    "user_id": row.user_id,
                    "corrections": row.corrections,
                    "avg_confidence": float(row.avg_confidence or 0),
                }
                for row in user_activity
            ],
            "period": {"start": start.isoformat(), "end": end.isoformat()},
        }

    def _gather_training_stats(self) -> Dict:
        """Gather model training statistics."""
        try:
            rows = self.db.execute(text("""
                SELECT version_name, cer_score, wer_score, trained_on_count,
                       deployed_at, is_active
                FROM model_versions
                ORDER BY deployed_at DESC
                LIMIT 10
            """)).fetchall()
            return {
                "model_versions": [
                    {
                        "version": row.version_name,
                        "cer": float(row.cer_score) if row.cer_score else None,
                        "wer": float(row.wer_score) if row.wer_score else None,
                        "samples": row.trained_on_count,
                        "deployed_at": str(row.deployed_at) if row.deployed_at else None,
                        "active": row.is_active,
                    }
                    for row in rows
                ]
            }
        except Exception:
            return {"model_versions": []}

    def _generate_pdf_report(self, stats: Dict, training_stats: Dict, config: ReportConfig) -> bytes:
        """Generate PDF report using reportlab."""
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch, cm
            from reportlab.platypus import (
                SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
            )
            from reportlab.lib.enums import TA_CENTER, TA_RIGHT
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
        except ImportError:
            logger.error("reportlab not installed. Run: pip install reportlab")
            return self._generate_text_report(stats, training_stats, config)

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=2 * cm,
            leftMargin=2 * cm,
            topMargin=2 * cm,
            bottomMargin=2 * cm,
        )

        styles = getSampleStyleSheet()
        styles.add(ParagraphStyle(
            name='ArabicTitle',
            fontName='Helvetica-Bold',
            fontSize=18,
            alignment=TA_CENTER,
            spaceAfter=20,
        ))
        styles.add(ParagraphStyle(
            name='SectionHeader',
            fontName='Helvetica-Bold',
            fontSize=14,
            spaceBefore=15,
            spaceAfter=8,
            textColor=colors.HexColor('#2563eb'),
        ))

        elements = []

        # Title
        elements.append(Paragraph(config.title, styles['ArabicTitle']))
        elements.append(Paragraph(
            f"Period: {config.period_start.strftime('%Y-%m-%d')} to "
            f"{config.period_end.strftime('%Y-%m-%d')}",
            styles['Normal']
        ))
        elements.append(Paragraph(
            f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
            styles['Normal']
        ))
        elements.append(Spacer(1, 20))

        # Summary Section
        s = stats["summary"]
        elements.append(Paragraph("Correction Summary", styles['SectionHeader']))

        summary_data = [
            ["Metric", "Value"],
            ["Total Regions", str(s["total_regions"])],
            ["Corrected", f"{s['corrected']} ({s['correction_rate']:.1f}%)"],
            ["Gold Standard", str(s["gold_standard"])],
            ["Pending Review", str(s["pending"])],
            ["Average Confidence", f"{s['avg_confidence']:.1%}"],
            ["Medical Terms", str(s["medical_terms"])],
            ["Arabic Words", str(s["arabic"])],
            ["Latin Words", str(s["latin"])],
            ["Mixed Script", str(s["mixed"])],
        ]

        summary_table = Table(summary_data, colWidths=[3 * inch, 2 * inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2563eb')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8fafc')]),
        ]))
        elements.append(summary_table)
        elements.append(Spacer(1, 20))

        # Top Corrected Terms
        if stats["top_corrected_terms"]:
            elements.append(Paragraph("Top Corrected Terms", styles['SectionHeader']))
            terms_data = [["Term", "Corrections"]]
            for term in stats["top_corrected_terms"][:10]:
                terms_data.append([term["term"], str(term["count"])])

            terms_table = Table(terms_data, colWidths=[3 * inch, 2 * inch])
            terms_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2563eb')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ]))
            elements.append(terms_table)
            elements.append(Spacer(1, 20))

        # Model Versions
        if training_stats.get("model_versions"):
            elements.append(PageBreak())
            elements.append(Paragraph("Model Training History", styles['SectionHeader']))
            model_data = [["Version", "CER", "WER", "Samples", "Active"]]
            for mv in training_stats["model_versions"]:
                active_mark = "Yes" if mv["active"] else ""
                model_data.append([
                    mv["version"] or "-",
                    f"{mv['cer']:.4f}" if mv["cer"] else "-",
                    f"{mv['wer']:.4f}" if mv["wer"] else "-",
                    str(mv["samples"]),
                    active_mark,
                ])

            model_table = Table(model_data, colWidths=[1.5*inch, 1*inch, 1*inch, 1*inch, 0.8*inch])
            model_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#16a34a')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ]))
            elements.append(model_table)

        doc.build(elements)
        buffer.seek(0)
        return buffer.getvalue()

    def _generate_excel_report(self, stats: Dict, training_stats: Dict, config: ReportConfig) -> bytes:
        """Generate Excel report using openpyxl."""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            logger.error("openpyxl not installed. Run: pip install openpyxl")
            return self._generate_text_report(stats, training_stats, config)

        wb = openpyxl.Workbook()

        # Summary sheet
        ws = wb.active
        ws.title = "Summary"

        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        ws["A1"] = config.title
        ws["A1"].font = Font(bold=True, size=16)
        ws["A2"] = f"Period: {config.period_start.strftime('%Y-%m-%d')} to {config.period_end.strftime('%Y-%m-%d')}"
        ws["A3"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"

        s = stats["summary"]
        metrics = [
            ("Total Regions", s["total_regions"]),
            ("Corrected", s["corrected"]),
            ("Correction Rate", f"{s['correction_rate']:.1f}%"),
            ("Gold Standard", s["gold_standard"]),
            ("Pending Review", s["pending"]),
            ("Average Confidence", f"{s['avg_confidence']:.1%}"),
            ("Medical Terms", s["medical_terms"]),
            ("Arabic Words", s["arabic"]),
            ("Latin Words", s["latin"]),
            ("Mixed Script", s["mixed"]),
        ]

        row = 5
        ws.cell(row=row, column=1, value="Metric").font = header_font
        ws.cell(row=row, column=1).fill = header_fill
        ws.cell(row=row, column=2, value="Value").font = header_font
        ws.cell(row=row, column=2).fill = header_fill

        for metric_name, metric_value in metrics:
            row += 1
            ws.cell(row=row, column=1, value=metric_name).border = thin_border
            ws.cell(row=row, column=2, value=metric_value).border = thin_border

        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 15

        # Top Terms sheet
        if stats["top_corrected_terms"]:
            ws2 = wb.create_sheet("Top Terms")
            ws2.cell(row=1, column=1, value="Term").font = header_font
            ws2.cell(row=1, column=1).fill = header_fill
            ws2.cell(row=1, column=2, value="Corrections").font = header_font
            ws2.cell(row=1, column=2).fill = header_fill

            for i, term in enumerate(stats["top_corrected_terms"], 2):
                ws2.cell(row=i, column=1, value=term["term"]).border = thin_border
                ws2.cell(row=i, column=2, value=term["count"]).border = thin_border

        # Model History sheet
        if training_stats.get("model_versions"):
            ws3 = wb.create_sheet("Model History")
            headers = ["Version", "CER", "WER", "Samples", "Deployed", "Active"]
            for col, header in enumerate(headers, 1):
                cell = ws3.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = PatternFill(start_color="16A34A", end_color="16A34A", fill_type="solid")

            for i, mv in enumerate(training_stats["model_versions"], 2):
                ws3.cell(row=i, column=1, value=mv.get("version", "")).border = thin_border
                ws3.cell(row=i, column=2, value=f"{mv['cer']:.4f}" if mv.get("cer") else "").border = thin_border
                ws3.cell(row=i, column=3, value=f"{mv['wer']:.4f}" if mv.get("wer") else "").border = thin_border
                ws3.cell(row=i, column=4, value=mv.get("samples", 0)).border = thin_border
                ws3.cell(row=i, column=5, value=mv.get("deployed_at", "")).border = thin_border
                ws3.cell(row=i, column=6, value="Yes" if mv.get("active") else "No").border = thin_border

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    def _generate_text_report(self, stats: Dict, training_stats: Dict, config: ReportConfig) -> bytes:
        """Fallback: Generate plain text report."""
        lines = [
            f"={config.title}=",
            f"Period: {config.period_start} to {config.period_end}",
            f"Generated: {datetime.now().isoformat()}",
            "",
            "Summary:",
        ]
        for key, value in stats["summary"].items():
            lines.append(f"  {key}: {value}")
        
        lines.append("\nTop Corrected Terms:")
        for term in stats["top_corrected_terms"][:10]:
            lines.append(f"  {term['term']}: {term['count']} corrections")
        
        return "\n".join(lines).encode("utf-8")
